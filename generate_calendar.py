#!/usr/bin/env python3

"""Downloads, extracts, and uploads schedules for AHS CZ pharmacists.
    Last Update: 2017-Mar-11

    Copyright (c) Notices
        2017  Joshua R. Torrance  <studybuffalo@studybuffalo.com>
    
    This program is free software: you can redistribute it and/or 
    modify it under the terms of the GNU General Public License as 
    published by the Free Software Foundation, either version 3 of the 
    License, or (at your option) any later version.
    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.
    You should have received a copy of the GNU General Public License
    along with this program.  If not, 
    see <http://www.gnu.org/licenses/>.
    SHOULD YOU REQUIRE ANY EXCEPTIONS TO THIS LICENSE, PLEASE CONTACT 
    THE COPYRIGHT HOLDERS.
"""

"""STYLE RULES FOR THIS PROGRAM
    Style follows the Python Style Guide (PEP 8) where possible. The 
    following are common standards for reference
    
    COMMENT LINES to max of 72 characters
    PROGRAM LINES to a max of 79 characters
    
    INDENTATION 4 spaces
    STRINGS use quotation marks
    VARIABLES use camelCase
    GLOBAL VARIABLES use lowercase with underscores
    CLASSES use CapWords
    CONSTANTS use UPPERCASE
    FUNCTIONS use lowercase with underscores
    MODULES use lowercase with underscores
    
    ALIGNMENT
        If possible, align with open delimiter
        If not possible, indent
        If one indent would align arguments with code in block, use 
            two indents to provide visual differentiation
        Operators should occur at start of line in broken up lines, 
        not at the end of the preceding line
    OPERATORS & SPACING
    Use spacing in equations
        e.g. 1 + 1 = 2
    Do not use spacing in assigning arguments in functions 
        e.g. def foo(bar=1):
"""

# IMPORTS
import sys
from unipath import Path
import python_logging
import requests
from requests_ntlm import HttpNtlmAuth
from datetime import datetime, timedelta, date
import csv
import openpyxl
import xlrd
import configparser
import pymysql
import re
import ftplib
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# CLASSES
class User(object):
    """Holds retrieved user data from the MySQL database"""

    def __init__(self, name, email, codeName, full, role, public, start):
        self.name = name
        self.email = email
        self.codeName = codeName
        self.fullDay = True if full == 1 else False
        self.role = role
        self.public = True if public == 1 else False
        self.index = 0
        self.start = start

class ShiftCode(object):
    """Holds the retrieved shift code details from MySQL database"""

    def __init__(self, code, mStart, mDuration, tuStart, tuDuration, wStart, 
                 wDuration, thStart, thDuration, fStart, fDuration, saStart, 
                 saDuration, suStart, suDuration, stStart, stDuration):
        self.code = code
        self.mStart = mStart
        self.mDuration = mDuration
        self.tuStart = tuStart
        self.tuDuration = tuDuration
        self.wStart = wStart
        self.wDuration = wDuration
        self.thStart = thStart
        self.thDuration = thDuration
        self.fStart = fStart
        self.fDuration = fDuration
        self.saStart = saStart
        self.saDuration = saDuration
        self.suStart = suStart
        self.suDuration = suDuration
        self.stStart = stStart
        self.stDuration = stDuration

class Shift(object):
    """Holds the details for a user's specified shift details"""
    
    def __init__(self, shift, date, comment):
        self.shift = shift
        self.startDate = date
        self.comment = comment

class Schedule(object):
    """Holds all the users shifts and any noted modifications"""

    def __init__(self, shifts, additions, deletions, changes, missing, null):
        self.shifts = shifts
        self.additions = additions
        self.deletions = deletions
        self.changes = changes
        self.missing = missing
        self.null = null

class EmailShift(object):
    """Holds details on shift modifications for emailing to the user"""
    date = 0
    shift = ""
    msg = ""

    def __init__(self, date, shift, msg):
        self.date = date
        self.shift = shift
        self.msg = msg


# FUNCTIONS
def get_date():
    """Generates todays date as string (in format yyyy-mm-dd)"""
    today = datetime.today()
    year = today.year
    month = "%02d" % today.month
    day = "%02d" % today.day
    date = "%s-%s-%s" % (year, month, day)
    
    return date

def database_connect():
    try:
        db = priCon.get('db', 'db')
        host = priCon.get('db', 'host')
        dbUser = priCon.get('db', 'user')
        dbPass = priCon.get('db', 'password')
    except:
        log.exception("Error accessing database credentials")

    try:
        conn = pymysql.connect(host, dbUser, dbPass, db)
        cursor = conn.cursor()
    except:
        log.exception("Unable to connect to database")
        cursor = None

    return cursor

def stat_holidays():
    """Reads the holidays.csv file to flag AHS Statutory Holidays
    
    Location of the holiday.csv file is managed by the config.cfg file

    CSV File Format
        Each line represents one holiday date
        Dates are represented as an integer (zero-leading not required)
        Format: YYYY, mm, dd

    AHS Statutory Holidays (as of last program update)
        New Years Day (Jan 1)
        Family Day (Third Monday of February)
        Good Friday (Friday before Easter)
        Victoria Day (Monday before May 25)
        Canada Day (July 1)
        Heritage Day (First Monday of August)
        Labour Day (First Monday of September)
        Thanksgiving (Second Monday of October)
        Rememberance Day (Novemer 11)
        Christmas (Dec 25)
        Boxing Day (Dec 26)
    """
    
    dates = {};

    csvPath = root.child("holidays.csv").absolute()

    try:
        with open(csvPath, "r") as csvFile:
            file = csv.reader(csvFile, delimiter=",", quotechar='"')

            for line in file:
                # Convert CSV strings to integers
                year = int(line[0])
                month = int(line[1])
                day = int(line[2])
            
                # Adds date to list
                if year in dates:
                    dates[year].append(date(year, month, day))
                else:
                    # Checks if year exists, if not adds it 
                    dates[year] = []
                    dates[year].append(date(year, month, day))
    except:
        log.exception("Error reading holiday dates from holiday.csv")

    return dates



def get_users(cursor):
    """Retrieves a list of all users in the specified MySQL database"""

    output = []

    # Checks if this should run in debug mode or not
    if pubCon.getboolean("debug", "program"):
        # Creates a user based on the config file
        output.append(User(
            pubCon.get("debug", "name"),
            pubCon.get("debug", "email"),
            pubCon.get("debug", "code_name"),
            pubCon.getint("debug", "full_day"),
            pubCon.get("debug", "role"),
            pubCon.getint("debug", "public"),
            datetime.strptime(pubCon.get("debug", "start_date"), "%Y-%m-%d"),
        ))

        log.info("Skipped user retrieval, running in debug mode")
    else:
        # Selects all user from the database
        query = ("SELECT user, email, code_name, full_day, role, public, "
                 "start_date FROM calendar_users")

        cursor.execute(query)
        
        for row in cursor:
            output.append(User(row[0], row[1], row[2], row[3], row[4], row[5], row[6]))

        log.info("Users retrieve successfully from database table")

    return output

def get_formatted_date(date):
    """Converts Python date object into string (as yyyy-mmm-dd)"""
    day = date.strftime("%d")
    
    month = date.strftime("%m")
    
    if month == "01":
        month = "JAN"
    elif month == "02":
        month = "FEB"
    elif month == "03":
        month = "MAR"
    elif month == "04":
        month = "APR"
    elif month == "05":
        month = "MAY"
    elif month == "06":
        month = "JUN"
    elif month == "07":
        month = "JUL"
    elif month == "08":
        month = "AUG"
    elif month == "09":
        month = "SEP"
    elif month == "10":
        month = "OCT"
    elif month == "11":
        month = "NOV"
    elif month == "12":
        month = "DEC"
    
    year = date.strftime("%Y")
    
    return ("%s-%s-%s" % (year, month, day))

def return_column_index(sheet, user):
    """Determines the Excel column containing the provided user"""

    if user.role == "a":
        nameRow = pubCon.getint("schedules", "name_row_a")
        nameColStart = pubCon.getint("schedules", "name_col_start_a")
        nameColEnd = pubCon.getint("schedules", "name_col_end_a")
    elif user.role == "p":
        nameRow = pubCon.getint("schedules", "name_row_p")
        nameColStart = pubCon.getint("schedules", "name_col_start_p")
        nameColEnd = pubCon.getint("schedules", "name_col_end_p")
    elif user.role == "t":
        nameRow = pubCon.getint("schedules", "name_row_t")
        nameColStart = pubCon.getint("schedules", "name_col_start_t")
        nameColEnd = pubCon.getint("schedules", "name_col_end_t")
    else:
        log.warn("User %s has invalid role: %s" % (user.name, user.role))

    # Cycles through names in Excel document to find the user's column
    index = None

    for i in range(nameColStart, nameColEnd):
        try:
            if user.role == "p":
                cellName = str(sheet.cell(row=nameRow, column=i).value).strip()
            elif user.role == "a" or user.role == "t":
                cellName = str(sheet.cell(nameRow, i).value).strip()
            
            if cellName == user.name:
                index = i
                break
        except IndexError:
            break
        except Exception:
            log.exception("Error while searcing for column index for %s" 
                          % user.name)

    return index

def compile_schedule_details(user, book, sheet, cursor):
    """Assembles users schedule and any changes
    
    Args:
        user: a user object
        book: the excel workbook containing the user's schedule
        sheet: the excel worksheet containing the user's schedule
        cursor:a pymysql object connected to the calendar database
    
    Returns:
        Returns a Schedule object
        
    Raises:
        None.
    """

    def collect_shift_times(cur, name, role):
        """Takes a specific user and extracts the shift times"""

        query = ("SELECT shift, m_start, m_duration, tu_start, tu_duration, "
                 "w_start, w_duration, th_start, th_duration, f_start, "
                 "f_duration, sa_start, sa_duration, su_start, su_duration, "
                 "st_start, st_duration FROM calendar_shifts "
                 "WHERE user = %s and role = %s")

        # Collect default shift codes
        args = ("DEFAULT", role)

        try:
            cur.execute(query, args)
        except:
            log.exception("Unable to retrieve default shift times for %s" 
                          % name)
        
        defaultCodes = []
    
        for row in cur:
            code = ShiftCode(row[0], row[1], row[2], row[3], row[4], row[5], 
                             row[6], row[7], row[8], row[9], row[10], 
                             row[11], row[12], row[13], row[14], row[15], 
                             row[16])
            defaultCodes.append(code)
    
        # Collect user-specific shift codes
        args = (name, role)
        try:
            cur.execute(query, args)
        except:
            log.exception("Unable to retrieve user shift times for %s" 
                          % name)

        userCodes = []
    
        for row in cur:
            code = ShiftCode(row[0], row[1], row[2], row[3], row[4], row[5], 
                             row[6], row[7], row[8], row[9], row[10], 
                             row[11], row[12], row[13], row[14], row[15], 
                             row[16])
            userCodes.append(code)

        # Combine default shift codes and user-specific shift codes
        codeList = []
    
        for defaultCode in defaultCodes:
            override = False
        
            for userCode in userCodes:
                # If shift code is a duplicate, userCode takes precedence
                if userCode.code == defaultCode.code:
                    codeList.append(userCode)
                    override = True
                
            if override == False:
                codeList.append(defaultCode)
    
        # Add any shift codes unique to user and not in default
        for userCode in userCodes:
            unique = True
        
            for defaultCode in defaultCodes:
                if userCode.code == defaultCode.code:
                    unique = False
        
            if unique == True:
                codeList.append(userCode)

        return codeList
    
    def retrieve_old_schedule(cur, name, role):
        """Retrieves the user's previous schedule from the database"""
        query = ("SELECT date, shift FROM calendar_schedules "
                 "WHERE user = %s AND role = %s")
        args = (name, role)

        try:
            cur.execute(query, args)
        except:
            log.exception("Unable to retrieve %s's old schedule" % user)

        oldSchedule = []

        for row in cur:
            shift = Shift(row[1], row[0], "")
            oldSchedule.append(shift)
        
        return oldSchedule

    def is_stat(date):
        """Determines if the date is a stat holiday or not"""

        for holiday in holidays[date.year]:
            if date == holiday:
                return True
                break
        
        return False


    # EXTRACT SCHEDULE DETAILS FROM EXCEL DOCUMENT
    log.info("Extracting schedule details for %s" % user.name)

    # Collect the details of where the schedule details are located
    if user.role == "a":
        rowStart = pubCon.getint("schedules", "shift_row_start_a")
        rowEnd = pubCon.getint("schedules", "shift_row_end_a")
        dateCol = pubCon.getint("schedules", "date_col_a")
    elif user.role == "p":
        rowStart = pubCon.getint("schedules", "shift_row_start_p")
        rowEnd = pubCon.getint("schedules", "shift_row_end_p")
        dateCol = pubCon.getint("schedules", "date_col_p")
    elif user.role == "t":
        rowStart = pubCon.getint("schedules", "shift_row_start_t")
        rowEnd = pubCon.getint("schedules", "shift_row_end_t")
        dateCol = pubCon.getint("schedules", "date_col_t")
    
    shifts = []

    # Generate comment map if this is an xls file
    if user.role == "a" or user.role == "t":
        commentMap = sheet.cell_note_map

    # Cycle through each row and extract shift date, code, and comments
    for i in range(rowStart, rowEnd):
        # Extract date
        try:
            if user.role == "p":
                date = sheet.cell(row=i, column=dateCol).value.date()
            elif user.role == "a" or user.role == "t":
                date = xlrd.xldate_as_tuple(
                    sheet.cell(i, dateCol).value, book.datemode
                )
                date = datetime(*date).date()
        except:
            date = ""

		# Extract shift code
        try:
            if user.role == "p":
                shiftCodes = sheet.cell(row=i, column=user.index).value.upper()
            elif user.role == "a" or user.role == "t":
                shiftCodes = sheet.cell(i, user.index).value.upper()
        except:
            shiftCodes = ""

        # Extract cell comments
        comment = ""

        try:
            if user.role == "p":
                comment = sheet.cell(row=i, column=user.index).comment
            elif user.role == "a" or user.role == "t":
                comment = commentMap[i, user.index].text

            if comment is None:
                # Replaces "None" comments as empty string for calendar use
                comment = ""
            else:
                comment = str(comment)
                comment = comment.replace("\n", " ")
                comment = comment.strip()
        except:
            comment = ""
        
        # Add shift to master list if it has a date and shift code
        if shiftCodes != "" and date != "":
            shiftCodes = re.split("\s+|/", shiftCodes)
            
            for code in shiftCodes:
                shifts.append(Shift(code, date, comment))
    
    
    # ASSIGN TIMES TO EXTRACTED SHIFTS
    # Get shift codes/times for user
    codeList = collect_shift_times(cursor, user.name, user.role)

    # Assign start and end date/times to user's shifts
    schedule = []
    nullShifts = []
    missingShifts = []
    
    for day in shifts:
        # Checks if this is a stat holiday and/or weekend
        statHoliday = is_stat(day.startDate)
        dow = day.startDate.weekday()
        
        # If shift is call, special processing required
        if day.shift == "CALL":
            day.shift = "Call"
            
            if day.startDate.weekday() < 5 and statHoliday == False:
                # Weekday call starts at 22:00
                day.startTime = timedelta(hours=22)
            else:
                # Weekend/Stat call starts at 19:30
                day.startTime = timedelta(hours=19, minutes=30)
            
            endDate = day.startDate + timedelta(days=1)
            day.endDate = endDate

            day.endTime = timedelta(hours=7)

            schedule.append(day)
        else:
            # Search for a shift match
            shiftMatch = False
            
            for code in codeList:
	            # If matched, find the proper day to base shift details on
                if day.shift == code.code:
                    shiftMatch = True
                    
                    if statHoliday == True:
                        startTime = code.stStart
                        duration = code.stDuration
                    elif dow == 0:
                        startTime = code.mStart
                        duration = code.mDuration
                    elif dow == 1:
                        startTime = code.tuStart
                        duration = code.tuDuration
                    elif dow == 2:
                        startTime = code.wStart
                        duration = code.wDuration
                    elif dow == 3:
                        startTime = code.thStart
                        duration = code.thDuration
                    elif dow == 4:
                        startTime = code.fStart
                        duration = code.fDuration
                    elif dow == 5:
                        startTime = code.saStart
                        duration = code.saDuration
                    elif dow == 6:
                        startTime = code.suStart
                        duration = code.suDuration

                    if startTime:
                        day.startTime = startTime
                        day.endDate = day.startDate

                        endTime = (startTime + duration)

                        # Adjusting times overlapping multiple days
                        if (endTime > timedelta(hours=24)):
                            days = timedelta(endTime.days)
                            day.endDate = day.endDate + days
                            endTime = endTime - days

                        day.endTime = endTime
                        
                        schedule.append(day)
                    else:
                        # Shift has no times - add it to the Null shift list
                        msg = "%s - %s" % (get_formatted_date(day.startDate), day.shift)
                        nullShifts.append(EmailShift(day.startDate, day.shift, msg))

            # If no shift match, provide default values
            if shiftMatch == False:
                # Add missing shift to the Missing shift list
                msg = "%s - %s" % (get_formatted_date(day.startDate), day.shift)
                missingShifts.append(EmailShift(day.startDate, day.shift, msg))
                
                # Set default times
                day.startTime = timedelta(hours=7)
                day.endDate = day.startDate
                
                if dow < 5 and statHoliday == False:
                    day.endTime = timedelta(hours=22)
                else:
                    day.endTime = timedelta(hours=19, minutes=30)

                schedule.append(day)

    
    # Determine the shift additions, deletions, and changes
    # Retrieve the old schedule
    oldSchedule = retrieve_old_schedule(cursor, user.name, user.role)
    
    # Check if there are any deletions or changes
    deletions = []
    changes = []

    for oldShift in oldSchedule:
        shiftDelete = True
        shiftChange = []
        changedShifts = []

        for newShift in schedule:
            if oldShift.startDate == newShift.startDate:
                shiftDelete = False

                if oldShift.shift == newShift.shift:
                    shiftChange.append(False)
                else:
                    shiftChange.append(True)
                    changedShifts.append(newShift.shift)

        # Adds deletions
        if shiftDelete == True:
            date = get_formatted_date(oldShift.startDate)
            msg = "%s - %s" % (get_formatted_date(oldShift.startDate), oldShift.shift)
            deletions.append(EmailShift(oldShift.startDate, oldShift.shift, msg))

        # Adds changes
        if len(shiftChange) and False not in shiftChange:
            shifts = "/".join(map(str, changedShifts))
            msg = "%s - %s changed to %s" % (get_formatted_date(oldShift.startDate), oldShift.shift, shifts)
            changes.append(EmailShift(oldShift.startDate, shifts, msg))

    # Checks if there are any new additions
    additions = []

    for newShift in schedule:
        shiftAdd = True

        for oldShift in oldSchedule:
            if oldShift.startDate == newShift.startDate:
                shiftAdd = False

        if shiftAdd == True:
            msg = "%s - %s" % (get_formatted_date(newShift.startDate), newShift.shift,)
            additions.append(EmailShift(newShift.startDate, newShift.shift, msg))

    # Checks if any of the null or missing shifts are new (i.e. haven't 
    # been reported to user yet)
    # Collects dates that could be new (additions or changes)
    modDate = []

    for a in additions:
        modDate.append(a.date)

    for d in deletions:
        modDate.append(d.date)

    for c in changes:
        modDate.append(c.date)

    # Removes any missing or null shifts not in the mod dates or are blank
    missing = []
    null = []

    for date in modDate:
        for shift in missingShifts:
            if shift.date == date and shift.shift.strip() != "":
                missing.append(shift)

        for shift in nullShifts:
            if shift.date == date and shift.shift.strip() != "":
                null.append(shift)

    return Schedule(schedule, additions, deletions, changes, missing, null)

def generate_calendar(schedule, user):
    """Generates an .ics file from the extracted user schedule"""

    log.info("Generating .ics calendar for %s" % user.name)

    # Generate initial calendar information
    lines = []
    
    lines.append("BEGIN:VCALENDAR")
    lines.append("PRODID:-//StudyBuffalo.com//RDRHC Calendar//EN")
    lines.append("VERSION:2.0")
    lines.append("CALSCALE:GREGORIAN")
    lines.append("X-WR-CALNAME:Work Schedule")
    lines.append("X-WR-TIMEZONE:America/Edmonton")
    lines.append("BEGIN:VTIMEZONE")
    lines.append("TZID:America/Edmonton")
    lines.append("X-LIC-LOCATION:America/Edmonton")
    lines.append("BEGIN:DAYLIGHT")
    lines.append("TZOFFSETFROM:-0700")
    lines.append("TZOFFSETTO:-0600")
    lines.append("TZNAME:MDT")
    lines.append("DTSTART:19700308T020000")
    lines.append("RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=2SU")
    lines.append("END:DAYLIGHT")
    lines.append("BEGIN:STANDARD")
    lines.append("TZOFFSETFROM:-0600")
    lines.append("TZOFFSETTO:-0700")
    lines.append("TZNAME:MST")
    lines.append("DTSTART:19701101T020000")
    lines.append("RRULE:FREQ=YEARLY;BYMONTH=11;BYDAY=1SU")
    lines.append("END:STANDARD")
    lines.append("END:VTIMEZONE")
    
    
	# Cycle through schedule and generate schedule events
    dtstamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    i = 0

    for day in schedule.shifts:
        startDate = day.startDate.strftime("%Y%m%d")
        comment = day.comment

        lines.append("BEGIN:VEVENT")

        if user.fullDay == False:	
            startTime = str(day.startTime).replace(":", "").zfill(6)
            endDate = day.endDate.strftime("%Y%m%d")
            endTime = str(day.endTime).replace(":", "").zfill(6)

            lines.append("DTSTART;TZID=\"America/Edmonton\":%sT%s" % (startDate, startTime))
            lines.append("DTEND;TZID=\"America/Edmonton\":%sT%s" % (endDate, endTime))
        elif user.fullDay == True:
            startTime = "000000"
            endDate = day.startDate + timedelta(days=1)
            endDate = endDate.strftime("%Y%m%d")

            lines.append("DTSTART;VALUE=DATE:%s" % startDate)
            lines.append("DTEND;VALUE=DATE:%s" % endDate)
            # DTSTART;VALUE=DATE:20170731
            # DTEND;VALUE=DATE:20170801
        
        lines.append("DTSTAMP:%s" % dtstamp)
        lines.append("UID:%sT%s@studybuffalo.com-%s" % (startDate, startTime, i))
        lines.append("CREATED:%s" % dtstamp)
        lines.append("DESCRIPTION:%s" % comment)
        lines.append("LAST-MODIFIED:%s" % dtstamp)
        lines.append("LOCATION:Red Deer Regional Hospital Centre")
        lines.append("SEQUENCE:0")
        lines.append("STATUS:CONFIRMED")
        lines.append("SUMMARY:%s Shift" % day.shift)
        lines.append("TRANSP:TRANSPARENT")
        lines.append("END:VEVENT")

        i = i + 1


    # End calendar file
    lines.append("END:VCALENDAR")

    # Fold any lines > 75 characters long
    folded = []
    
    for line in lines:
        if len(line) > 75:

            # Create line with first 75 characters
            newLine = line[0:75]
            folded.append(newLine + "\n")

            # Go through remainder and fold them
            line = line[75:]
            length = len(line)

            while length > 75:
                # Add folded line
                newLine = line[0:75]
                folded.append(" " + newLine + "\n")
                
                # Generate new line and length
                line = line[75:]
                length = len(line)

            # Add remainder
            folded.append(" " + line + "\n")
        else:
            folded.append(line + "\n")

    # Cycle through schedule list and generate .ics file
    fileName = user.codeName
    schedTitle = "%s - %s.ics" % (today, fileName)
    fileLoc = root.child("calendars", schedTitle).absolute()

    with open(fileLoc, "w") as ics:
        for line in folded:
            ics.write(line)

def upload_calendar(user):
    """Uploads .ics calendar to server"""
    address = priCon.get("ftp", "address")
    username = priCon.get("ftp", "user")
    password = priCon.get("ftp", "password")
    
    ftp = ftplib.FTP(address, username, password)

    # Open the calendar file
    fileName = user.codeName
    schedTitle = "%s - %s.ics" % (today, fileName)
    fileLoc = root.child("calendars", schedTitle).absolute()
    calFile = open(fileLoc, 'rb')

    # Generate website name and STOR message
    onlineName = "%s.ics" % user.codeName

    storMsg = "STOR %s" % onlineName

    # Chooses location based on public or private access
    if user.public is True:
        ftpLoc = pubCon.get("ftp", "public_location")
    elif user.public is False:
        ftpLoc = pubCon.get("ftp", "private_location")
    
    # Set FTP directory
    try:
        ftp.cwd(ftpLoc)
    except:
        log.exception("Unable to change FTP directory to %s" % ftpLoc)
    
    # Upload the .ics calendar to the server
    try:
        log.info("Uploading .ics calendar to server for %s" % user.name)

        ftp.storlines(storMsg, calFile)
    
        calFile.close()
        ftp.quit()
    except:
        log.exception("Unable to upload .ics calendar to server for %s" 
                      % user.name)

def update_db(schedule, cursor):
    """Uploads user schedule to MySQL database"""

    # Delete all items for this user
    query = "DELETE FROM calendar_schedules WHERE user = %s AND role = %s"
    args = (user.name, user.role)

    try:
        cursor.execute(query, args)
    except:
        log.exception("Unable to remove old schedule from database for %s" 
                      % user.name)
    
    # Cycle through and create an array of mysql data
    mysqlData = []
    
    for day in schedule.shifts:
        mysqlData.append((day.startDate, day.shift, user.name, user.role))
    
    # Upload the new schedule
    query = ("INSERT INTO calendar_schedules (date, shift, user, role) "
             "VALUES (%s, %s, %s, %s)")

    try:
        cursor.executemany(query, mysqlData)
    except:
        log.exception("Unable to upload new schedule to database for %s" 
                      % user.name)

def email_welcome(user):
    """Sends a welcome email to any new user"""

    # Get dates of the user's sign up
    end = datetime.now()
    start = end - timedelta(days=1)
    
    # Check if user needs email sent (signed up in past 24 hours)
    try:
        if user.start > start and user.start < end:
            sendEmail = True
        else:
            sendEmail = False
    except:
        sendEmail = False

        log.exception("Unable to check %s's start date" % user.name)

    # Send email if user started in past 24 hours
    if sendEmail:
        server = priCon.get("email", "server")
        login = priCon.get("email", "user")
        pw = priCon.get("email", "password")
        fromEmail = login
        toEmail = user.email
        subject = "Welcome to Your New Online Schedule"

        content = MIMEMultipart('alternative')
        content['From'] = fromEmail
        content['To'] = toEmail
        content['Subject'] = subject

        # Collects text welcome email from template file
        textLoc = pubCon.get("email", "welcome_text", raw=True)
        
        try:
            with open(textLoc, "r") as textFile:
                text = textFile.read().replace("\n", "\r\n")
        except:
            log.exception("Unable to read welcome email text template")
            return

        # Collects html welcome email from template file
        htmlLoc = pubCon.get("email", "welcome_html", raw=True)
            
        try:
            with open(htmlLoc, "r") as htmlFile:
                html = htmlFile.read()
        except:
            log.exception("Unable to read welcome email html template")
            return

        # Assemble an HTML and plain text version
        textBody = MIMEText(text, 'plain')
        htmlBody = MIMEText(html, 'html')
        
        content.attach(textBody)
        content.attach(htmlBody)
        
        # Attempt to send email
        try:
            log.info("Sending welcome email to %s" % user.name)

            server = smtplib.SMTP(server)
            server.ehlo()
            server.starttls()
            server.login(login, pw)
            server.sendmail(fromEmail, toEmail, content.as_string())
            server.quit()
        except:
            log.exception("Unable to send welcome email to %s" % user.name)

def email_schedule(schedule, user):
    """Emails user with any schedule changes"""

    if (len(schedule.additions) or len(schedule.deletions) or 
        len(schedule.changes) or len(schedule.missing) or len(schedule.null)):
        
        login = priCon.get('email', 'user')
        pw = priCon.get('email', 'password')
        fromEmail = login
        toEmail = user.email
        subject = "RDRHC Schedule Changes"
        
        content = MIMEMultipart('alternative')
        content['From'] = fromEmail
        content['To'] = toEmail
        content['Subject'] = subject
        
        text = []
        html = []
        
        # Initial HTML Setup
        html.append("<html>")
        html.append("<head></head>")
        html.append("<body>")
        
        # Email opening
        text.append("Hello %s,\r\n" % user.name)
        text.append("Please see the following details regarding your work "
                    "schedule at the Red Deer Regional Hospital Centre:\r\n")
        
        html.append("<p>Hello %s,</p>" % user.name)
        html.append("<p>Please see the following details regarding your "
                    "work schedule at the Red Deer Regional Hospital "
                    "Centre:</p>")
        
        # Cycle through changes and create html & text email messages
        if len(schedule.additions):
            text.append("ADDITIONS")
            text.append("------------------------------------")
            
            html.append("<b>ADDITIONS</b>")
            html.append("<ul>")
            
            for a in schedule.additions:
                text.append(a.msg)
                
                html.append("<li>%s</li>" % a.msg)
                
            text.append("")
            
            html.append("</ul>")
            
        if len(schedule.deletions):
            text.append("DELETIONS")
            text.append("------------------------------------")
            
            html.append("<b>DELETIONS</b>")
            html.append("<ul>")
            
            for d in schedule.deletions:
                text.append(d.msg)
                
                html.append("<li>%s</li>" % d.msg)
                
            text.append("")

            html.append("</ul>")
            
        if len(schedule.changes):
            text.append("CHANGES")
            text.append("------------------------------------")
            
            html.append("<b>CHANGES</b>")
            html.append("<ul>")
            
            for c in schedule.changes:
                text.append(c.msg)
                
                html.append("<li>%s</li>" % c.msg)
                
            text.append("")
            
            html.append("</ul>")
            
        if len(schedule.missing):
            text.append("MISSING SHIFT CODES")
            text.append("------------------------------------")
            text.append("A default shift time of 07:00 to 22:00 (weekdays) "
                        "or 07:00 to 19:30 (weekends and stats) has been "
                        "used for these shifts")
            
            html.append("<b>MISSING SHIFT CODES</b>")
            html.append("<br><i>A default shift time of 07:00 to 22:00 "
                        "(weekdays) or 07:00 to 19:30 (weekends and stats) "
                        "has been used for these shifts</i>")
            html.append("<ul>")
            
            for m in schedule.missing:
                text.append(m.msg)
                
                html.append("<li>%s</li>" % m.msg)
                
            text.append("")
            
            html.append("</ul>")
            
        if len(schedule.null):
            text.append("EXCLUDED CODES")
            text.append("------------------------------------")
            text.append("These codes are for you to review to ensure no "
			            "work shifts have been missed; these codes have " 
			            "interpretted either as holidays/vacations/sick " 
			            "time/etc. or as being unrelated to start and end " 
			            "times")
            
            html.append("<b>EXCLUDED CODES</b>")
            html.append("<br><i>These codes are for you to review to "
                        "ensure no work shifts have been missed; these "
                        "codes have interpretted either as holidays/"
                        "vacations/sick time/etc. or as being unrelated to "
                        "start and end times</i>")
            html.append("<ul>")
            
            for n in schedule.null:
                text.append(n.msg)
                
                html.append("<li>%s</li>" % n.msg)
                
            text.append("")
            
            html.append("</ul>")
            
        # Add email footer
        # Generate File Name
        fileName = user.codeName

        # Generate File Location
        if user.public is True:
            fileLoc = fileName
        elif user.public is False:
            fileLoc = "private/%s" % fileName


        text.append("The address for your schedule is: "
                    "http://www.studybuffalo.com/calendar/%s.ics" % fileLoc)
        text.append("")
        
        html.append("<hr>")
        html.append("<p>The address for your schedule is: "
                    "http://www.studybuffalo.com/calendar/%s.ics </p>" 
                    % fileLoc)
        
        # Add tutorials
        text.append("For help using the calendar file, please see the "
                    "tutorials located at: "
                    "http://www.studybuffalo.com/calendar/")
        text.append("")
        
        html.append("<p>For help using the calendar file, please see the "
                    "tutorials located at: "
                    "http://www.studybuffalo.com/calendar/ </p>")
        
        # Add final notice
        text.append("If you wish to have any modifications made to your "
                    "shift codes (including exlcuded or missing shift "
                    "codes), please contact the owner of this program")
        
        html.append("<p>If you wish to have any modifications made to your "
                    "shift codes (including exlcuded or missing shift "
                    "codes), please contact the owner of this program</p>")
        
		# HTML finishing
        html.append("</body>")
        html.append("</html>")

        # Assemble final html and txt email content
        text = "\r\n".join(text)
        html = "".join(html)
        
        textBody = MIMEText(text, 'plain')
        htmlBody = MIMEText(html, 'html')
        
        content.attach(textBody)
        content.attach(htmlBody)
        
        try:
            log.info("Sending update email to %s" % user.name)
            
            server = smtplib.SMTP('smtp.gmail.com:587')
            server.ehlo()
            server.starttls()
            server.login(login, pw)
            server.sendmail(fromEmail, toEmail, content.as_string())
            server.quit()
        except:
            log.exception("Unable to send update email to %s" % user.name)


# SETUP FUNCTIONS
# Set root for this program to allow absolute paths
root = Path(sys.argv[1])

# Connect to public config file
pubCon = configparser.ConfigParser()
pubCon.read(root.child("calendar_config.cfg").absolute())

# Connect to private config file
priCon = configparser.ConfigParser()
priCon.read(Path(pubCon.get("misc", "priv_config", raw=True)).absolute())

# Set up logger
log = python_logging.start(priCon)

log.info("STARTING RDRHC CALENDAR GENERATOR")

# Get todays date for file saving
today = get_date()

# Generate database connection and cursor
cursor = database_connect()

# Set up file with list of stat holidays
holidays = stat_holidays()

# Download the schedules
if pubCon.getboolean("debug", "download"):
    download_schedules()
else:
    log.info("Skipped schedule download, running in debug mode")

# Get user names to build schedules for
if cursor:
    users = get_users(cursor)
else:
    log.error("No valid database cursor")
    users = []
 
# Go through each user and extract and process their schedule
for user in users:
    # Set the Excel information to open schedule
    if user.role == "a":
        fileName = ("%s_assistant.%s" 
                     % (today, pubCon.get("schedules", "type_a")))
        fileLoc = root.child("schedules", fileName).absolute()
        sheet = "current schedule"
    elif user.role == "p":
        fileName = ("%s_pharmacist.%s" 
                     % (today, pubCon.get("schedules", "type_p")))
        fileLoc = root.child("schedules", fileName).absolute()
        sheet = "current"
    elif user.role == "t":
        fileName = ("%s_technician.%s" 
                     % (today, pubCon.get("schedules", "type_t")))
        fileLoc = root.child("schedules", fileName).absolute()
        sheet = "current"
        
    # Open proper Excel worksheet
    if user.role == "p":
        excelBook = openpyxl.load_workbook(fileLoc)
        excelSheet = excelBook[sheet]
    elif user.role == "a" or user.role == "t":
        excelBook = xlrd.open_workbook(fileLoc)
        excelSheet = excelBook.sheet_by_name(sheet)
        
    # Find column index for this user
    user.index = return_column_index(excelSheet, user)
    
    # If the user.index is found, can run rest of program
    if user.index:
        # Get schedule for this user
        schedule = compile_schedule_details(user, excelBook, excelSheet, cursor)
    
        # Generate and save calendar file
        generate_calendar(schedule, user)
    
        # Upload calendar to website
        upload_calendar(user)
    
        # Update database with new schedule
        update_db(schedule, cursor)
    
        # Email user a welcome email
        email_welcome(user)
    
        # Email user with any schedule changes
        email_schedule(schedule, user)
    else:
        log.warn("Unable to find %s (role = %s) in the specified document"
                 % (user.name, user.role))

log.info("CALENDAR GENERATION COMPLETE")
log.info("")