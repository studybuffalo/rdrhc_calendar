"""Functions to extract schedule details from Excel file."""
from datetime import datetime
import logging
import re

import openpyxl
import xlrd

from modules.custom_exceptions import ScheduleError


LOG = logging.getLogger(__name__)


def return_column_index(sheet, user, cfg):
    """Determines the Excel column containing the provided user"""
    LOG.debug('Looking for user index in Excel schedule')

    index = None
    role = user['role']

    for i in range(cfg['col_start'], cfg['col_end']):
        try:
            if cfg['ext'] == 'xlsx':
                cell_name = str(
                    sheet.cell(row=cfg['name_row'], column=i).value
                ).strip()
            elif cfg['ext'] == 'xls':
                cell_name = str(sheet.cell(cfg['name_row'], i).value).strip()

            if cell_name.upper() == user['schedule_name'].upper():
                index = i
                break
        except IndexError:
            # Expected error if loop exceeds Excel content
            break

    if index:
        return index

    # No index found - raise error
    raise ScheduleError(
        f'Unable to find index for {user["name"]} (role = {role})'
    )


def format_shift_details(shift_codes, date, comment, role):
    """Splits up multiple shift codes and assigns details."""
    shifts = []

    if all([
            shift_codes != '',
            re.match(r'\s+$', shift_codes) is None,
            date != ''
    ]):
        # Split each shift code on spaces or slashes
        shift_codes = re.split(r'(?:\s|/)+', shift_codes.strip())

        # Remove any duplicate shift codes
        shift_codes = list(set(shift_codes))

        for code in shift_codes:
            if code != '' and re.match(r'\s+$', code) is None:
                shifts.append({
                    'shift_code': code,
                    'start_date': date,
                    'comment': comment,
                })

                # Add pharmacist 'X' shifts
                if role == 'p' and code[-1:].upper() == 'X':
                    shifts.append({
                        'shift_code': 'X',
                        'start_date': date,
                        'comment': '',
                    })
    return shifts


def _extract_date(cfg, book, sheet, i):
    """Extracts the date for a row."""
    date = ''

    try:
        if cfg['ext'] == 'xlsx':
            date = sheet.cell(row=i, column=cfg['date_col']).value.date()
        elif cfg['ext'] == 'xls':
            date = xlrd.xldate_as_tuple(
                sheet.cell(i, cfg['date_col']).value, book.datemode
            )
            date = datetime(*date).date()
    except AttributeError:
        # Expected error when there is no date value
        pass
    except IndexError:
        # Expected error when there is no date value
        pass
    except TypeError:
        # Expected error when there is no date value
        pass

    return date


def _extract_shift_codes(cfg, sheet, index, i):
    """Extracts the shift code for a row."""
    shift_codes = ''

    try:
        if cfg['ext'] == 'xlsx':
            shift_codes = sheet.cell(row=i, column=index).value.upper()
        elif cfg['ext'] == 'xls':
            shift_codes = sheet.cell(i, index).value.upper()
    except AttributeError:
        # Expected error when there is no shift code value
        pass
    except IndexError:
        # Expect error when there is no shift code value
        pass

    return shift_codes


def _extract_comment(cfg, sheet, comment_map, index, i):
    """Extracts the comments for a row."""
    comment = ''

    try:
        if cfg['ext'] == 'xlsx':
            comment = sheet.cell(row=i, column=index).comment
        elif cfg['ext'] == 'xls':
            comment = comment_map[i, index].text

        if comment is None:
            # Replaces 'None' comments as empty string for calendar use
            comment = ''
        else:
            comment = str(comment)
            comment = comment.replace('\n', ' ')
            comment = comment.strip()
    except KeyError:
        # Expected error when there is no comment
        pass

    return comment


def extract_raw_schedule(book, sheet, user, index, cfg):
    """Returns a list of schedule_shift objects."""

    # Extracts schedule details from spreadsheet
    LOG.info('Extracting schedule details for %s', user['name'])
    role = user['role']

    # Generate comment map if this is an xls file
    if cfg['ext'] == 'xls':
        comment_map = sheet.cell_note_map
    else:
        comment_map = None

    # Cycle through each row and extract shift date, code, and comments
    LOG.debug('Cycling through rows of excel schedule')

    shifts = []

    for i in range(cfg['row_start'], cfg['row_end']):
        date = _extract_date(cfg, book, sheet, i)
        shift_codes = _extract_shift_codes(cfg, sheet, index, i)
        comment = _extract_comment(cfg, sheet, comment_map, index, i)

        # Format and add shifts to the master list
        shifts.extend(format_shift_details(shift_codes, date, comment, role))

    # Sort the shifts by date
    # Note: should occur automatically, but just in case
    LOG.debug('Sorting shifts by date')

    sorted_shifts = sorted(shifts, key=lambda s: s['start_date'])

    return sorted_shifts


def _open_worksheet(config, file_loc, sheet_name, role):
    """Opens up the workbook and worksheet for this user."""
    if config['ext'] == 'xlsx':
        try:
            excel_book = openpyxl.load_workbook(file_loc)
            excel_sheet = excel_book[sheet_name]

            return excel_book, excel_sheet
        except FileNotFoundError:
            # Expected error when workbook not found
            pass
        except KeyError:
            # Expected error when worksheet does not exist
            pass
    if config['ext'] == 'xls':
        try:
            excel_book = xlrd.open_workbook(file_loc)
            excel_sheet = excel_book.sheet_by_name(sheet_name)

            return excel_book, excel_sheet
        except FileNotFoundError:
            # Expected error when workbook not found
            pass
        except xlrd.XLRDError:
            # Expected error when worksheet does not exist
            pass

    # No workbook or worksheet found - raise error
    raise ScheduleError(
        f'Cannot open .{config["ext"]} file for user role = {role}: {file_loc}'
    )


def generate_raw_schedule(app_config, excel_files, user):
    """Returns a list of shift details for the specified user."""

    # Setup the required Excel details
    role = user['role']
    file_loc = excel_files[role]
    config = app_config[f'{role}_excel'.lower()]

    # Open the proper Excel worksheet
    LOG.debug('Opening the Excel worksheet')
    excel_sheet = None

    raw_schedule = []

    for sheet_name in config['sheet']:
        try:
            excel_book, excel_sheet = _open_worksheet(config, file_loc, sheet_name, role)
        except ScheduleError:
            # Expected error when worksheet is not valid
            # Fails silently and continues loop
            continue

        # Find column index for this user
        try:
            user_index = return_column_index(excel_sheet, user, config)
        except ScheduleError:
            LOG.info(
                'Unable to find user index for %s (role = %s).',
                user['schedule_name'],
                user['role'],
            )
            user_index = None

        if user_index:
            raw_schedule += extract_raw_schedule(
                excel_book, excel_sheet, user, user_index, config,
            )

    if not raw_schedule:
        LOG.warning(
            'No shifts found for user %s (role = %s)',
            user['schedule_name'],
            user['role'],
        )

    return raw_schedule
