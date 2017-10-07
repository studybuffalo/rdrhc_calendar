from datetime import datetime, timedelta
from django.db.models import Q
import logging
import openpyxl
import re
import xlrd

# Setup logger
log = logging.getLogger(__name__)

# CLASSES
class User(object):
    """Holds retrieved user data from the MySQL database"""

    def __init__(self, name, email, codeName, full, role, public, start):
        self.name = name
        self.email = email
        self.codeName = codeName
        self.fullDay = True if full == 1 else False
        self.role = role
        self.public = True if public == 1 else False
        self.index = 0
        self.start = start

class RawShift(object):
    """Holds the details for a user's specified shift details"""
    
    def __init__(self, shift, date, comment):
        self.shift_code = shift
        self.start_date = date
        self.comment = comment

    def __str__(self):
        return "{} ({})".format(self.shift, self.start_date)

class FormattedShift(object):
    """Holds expanded details on a user's specified shift"""

    def __init__(self, code, start_datetime, end_datetime, comment, django):
        self.shift_code = code
        self.start_datetime = start_datetime
        self.end_datetime = end_datetime
        self.comment = comment
        self.django_shift = django

    def __str__(self):
        return "{} ({} to {})".format(
            self.shift_code, self.start_datetime, self.end_datetime
        )

class Schedule(object):
    """Holds all the users shifts and any noted modifications"""

    def __init__(self, shifts, additions, deletions, changes, missing, null):
        self.shifts = shifts
        self.additions = additions
        self.deletions = deletions
        self.changes = changes
        self.missing = missing
        self.null = null

class EmailShift(object):
    """Holds details on shift modifications for emailing to the user"""
    date = 0
    shift = ""
    msg = ""

    def __init__(self, date, shift, msg):
        self.date = date
        self.shift = shift
        self.msg = msg


def get_formatted_date(date):
    """Converts Python date object into string (as yyyy-mmm-dd)"""
    day = date.strftime("%d")
    
    month = date.strftime("%m")
    
    if month == "01":
        month = "JAN"
    elif month == "02":
        month = "FEB"
    elif month == "03":
        month = "MAR"
    elif month == "04":
        month = "APR"
    elif month == "05":
        month = "MAY"
    elif month == "06":
        month = "JUN"
    elif month == "07":
        month = "JUL"
    elif month == "08":
        month = "AUG"
    elif month == "09":
        month = "SEP"
    elif month == "10":
        month = "OCT"
    elif month == "11":
        month = "NOV"
    elif month == "12":
        month = "DEC"
    
    year = date.strftime("%Y")
    
    return ("%s-%s-%s" % (year, month, day))

def generate_calendar(user, schedule, root):
    """Generates an .ics file from the extracted user schedule"""
    
    log.info("Generating .ics calendar for %s" % user.name)

    # Generate initial calendar information
    lines = []
    
    lines.append("BEGIN:VCALENDAR")
    lines.append("PRODID:-//StudyBuffalo.com//RDRHC Calendar//EN")
    lines.append("VERSION:2.0")
    lines.append("CALSCALE:GREGORIAN")
    lines.append("X-WR-CALNAME:Work Schedule")
    lines.append("X-WR-TIMEZONE:America/Edmonton")
    lines.append("BEGIN:VTIMEZONE")
    lines.append("TZID:America/Edmonton")
    lines.append("X-LIC-LOCATION:America/Edmonton")
    lines.append("BEGIN:DAYLIGHT")
    lines.append("TZOFFSETFROM:-0700")
    lines.append("TZOFFSETTO:-0600")
    lines.append("TZNAME:MDT")
    lines.append("DTSTART:19700308T020000")
    lines.append("RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=2SU")
    lines.append("END:DAYLIGHT")
    lines.append("BEGIN:STANDARD")
    lines.append("TZOFFSETFROM:-0600")
    lines.append("TZOFFSETTO:-0700")
    lines.append("TZNAME:MST")
    lines.append("DTSTART:19701101T020000")
    lines.append("RRULE:FREQ=YEARLY;BYMONTH=11;BYDAY=1SU")
    lines.append("END:STANDARD")
    lines.append("END:VTIMEZONE")
    
    
	# Cycle through schedule and generate schedule events
    dt_stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    i = 0

    for shift in schedule:
        start_date = shift.start_datetime.strftime("%Y%m%d")
        comment = shift.comment

        lines.append("BEGIN:VEVENT")

        if user.full_day == False:	
            start_time = str(shift.start_datetime.time()).replace(":", "").zfill(6)
            end_date = shift.end_datetime.strftime("%Y%m%d")
            end_time = str(shift.end_datetime.time()).replace(":", "").zfill(6)

            lines.append("DTSTART;TZID=America/Edmonton:%sT%s" % (start_date, start_time))
            lines.append("DTEND;TZID=America/Edmonton:%sT%s" % (end_date, end_time))
        elif user.full_day == True:
            start_time = "000000"
            end_date = shift.start_datetime.date() + timedelta(days=1)
            end_date = end_date.strftime("%Y%m%d")

            lines.append("DTSTART;VALUE=DATE:%s" % start_date)
            lines.append("DTEND;VALUE=DATE:%s" % end_date)
        
        lines.append("DTSTAMP:%s" % dt_stamp)
        lines.append("UID:%sT%s@studybuffalo.com-%s" % (start_date, start_time, i))
        lines.append("CREATED:%s" % dt_stamp)
        lines.append("DESCRIPTION:%s" % comment)
        lines.append("LAST-MODIFIED:%s" % dt_stamp)
        lines.append("LOCATION:Red Deer Regional Hospital Centre")
        lines.append("SEQUENCE:0")
        lines.append("STATUS:CONFIRMED")
        lines.append("SUMMARY:%s Shift" % shift.shift_code)
        lines.append("TRANSP:TRANSPARENT")
        lines.append("END:VEVENT")

        i = i + 1


    # End calendar file
    lines.append("END:VCALENDAR")

    # Fold any lines > 75 characters long
    folded = []
    
    for line in lines:
        if len(line) > 75:

            # Create line with first 75 characters
            newLine = line[0:75]
            folded.append(newLine + "\n")

            # Go through remainder and fold them
            line = line[75:]
            length = len(line)

            while length > 75:
                # Add folded line
                newLine = line[0:75]
                folded.append(" " + newLine + "\n")
                
                # Generate new line and length
                line = line[75:]
                length = len(line)

            # Add remainder
            folded.append(" " + line + "\n")
        else:
            folded.append(line + "\n")

    # Cycle through schedule list and generate .ics file
    file_name = user.calendar_name
    sched_title = "%s.ics" % file_name
    file_loc = root.child("calendars", sched_title).absolute()

    with open(file_loc, "w") as ics:
        for line in folded:
            ics.write(line)

def extract_raw_schedule(book, sheet, user, index, row_start, row_end, date_col):
    """Returns an array of schedule_shift objects"""
    
    # EXTRACT SCHEDULE DETAILS FROM EXCEL DOCUMENT
    log.info("Extracting schedule details for %s" % user.name)

    # Generate comment map if this is an xls file
    if user.role == "a" or user.role == "t":
        comment_map = sheet.cell_note_map

    # Cycle through each row and extract shift date, code, and comments
    shifts = []

    for i in range(row_start, row_end):
        # Extract date
        try:
            if user.role == "p":
                date = sheet.cell(row=i, column=date_col).value.date()
            elif user.role == "a" or user.role == "t":
                date = xlrd.xldate_as_tuple(
                    sheet.cell(i, date_col).value, book.datemode
                )
                date = datetime(*date).date()
        except AttributeError:
            # Expected error when there is no date value
            date = ""
        except Exception as e:
            log.debug("Unable to extract date from worksheet")
            date = ""

		# Extract shift code
        try:
            if user.role == "p":
                shift_codes = sheet.cell(row=i, column=index).value.upper()
            elif user.role == "a" or user.role == "t":
                shift_codes = sheet.cell(i, index).value.upper()
        except AttributeError:
            # Expected error when there is no date value
            shift_codes = ""
        except Exception as e:
            log.debug("Unable to extract shift code from worksheet")
            shift_codes = ""

        # Extract cell comments
        comment = ""

        try:
            if user.role == "p":
                comment = sheet.cell(row=i, column=index).comment
            elif user.role == "a" or user.role == "t":
                comment = comment_map[i, index].text

            if comment is None:
                # Replaces "None" comments as empty string for calendar use
                comment = ""
            else:
                comment = str(comment)
                comment = comment.replace("\n", " ")
                comment = comment.strip()
        except:
            log.debug("Unable to extract comments from worksheet")
            comment = ""
        
        # Add shift to master list if it has a date and shift code
        if shift_codes != "" and date != "":
            # Split each shift code on spaces or slashes
            shift_codes = re.split("\s+|/", shift_codes)
            
            for code in shift_codes:
                shifts.append(RawShift(code, date, comment))
                
                # Add pharmacist "X" shifts
                if user.role == "p" and code[-1:].upper() == "X":
                    shifts.append(RawShift("X", date, ""))
    
    # Sort the shifts by date
    # Note: should occur automatically, but just in case
    sorted_shifts = sorted(shifts, key=lambda s: s.start_date)

    return sorted_shifts
    
def generate_formatted_schedule(user, raw_schedule, ShiftCode, StatHoliday, config, Shift):
    """Takes the raw schedule and returns the required formatted objects"""
    
    def collect_shift_codes(user, ShiftCode):
        """Takes a specific user and extracts the shift times"""
        # Collect the user-specific codes
        user_codes = ShiftCode.objects.filter(
            Q(role=user.role) & Q(user__exact=user.id)
        )

        # Collect the default codes (i.e. no user)
        default_codes = ShiftCode.objects.filter(
            Q(role=user.role) & Q(user__isnull=True)
        )

        # Add all the user_codes into the codes list
        codes = []
        
        for u_code in user_codes:
            codes.append(u_code)
        
        # Add any default codes that don't have a user code already
        for d_code in default_codes:
            if not any(d_code.code == code.code for code in codes):
                codes.append(d_code)
        
        return codes

    def collect_stat_holidays(schedule, StatHoliday):
        """Collects all stat holidays needed to generate a schedule"""
        first_day = schedule[0].start_date
        last_day = schedule[-1].start_date

        stat_holidays = StatHoliday.objects.all().filter(
            Q(date__gte=first_day) & Q(date__lte=last_day)
        )

        return stat_holidays

    def collect_defaults(config):
        weekday_start = datetime.strptime(
        config.get("schedules", "default_weekday_start"),
        "%H:%M:%S"
        ).time()

        weekend_start = datetime.strptime(
            config.get("schedules", "default_weekend_start"),
            "%H:%M:%S"
        ).time()

        stat_start = datetime.strptime(
            config.get("schedules", "default_stat_start"),
            "%H:%M:%S"
        ).time()

        weekday_duration = config.getfloat("schedules", "default_weekday_duration")
        weekday_hours = int(weekday_duration)
        weekday_minutes = int((weekday_duration*60) % 60)
        weekday_seconds = int((weekday_minutes*3600) % 3600)

        weekend_duration = config.getfloat("schedules", "default_weekend_duration")
        weekend_hours = int(weekend_duration)
        weekend_minutes = int((weekend_duration*60) % 60)
        weekend_seconds = int((weekend_minutes*60) % 60)

        stat_duration = config.getfloat("schedules", "default_stat_duration")
        stat_hours = int(stat_duration)
        stat_minutes = int((stat_duration*60) % 60)
        stat_seconds = int((stat_minutes*60) % 60)

        return {
            "weekday_start": weekday_start,
            "weekday_hours": weekday_hours,
            "weekday_minutes": weekday_minutes,
            "weekday_seconds": weekday_seconds,
            "weekend_start": weekday_start,
            "weekend_hours": weekday_hours,
            "weekend_minutes": weekday_minutes,
            "weekend_seconds": weekday_seconds,
            "stat_start": weekday_start,
            "stat_hours": weekday_hours,
            "stat_minutes": weekday_minutes,
            "stat_seconds": weekday_seconds,
        }

    def is_stat(stat_holidays, date):
        """Determines if the date is a stat holiday or not"""

        for holiday in stat_holidays:
            if date == holiday.date:
                return True
                break
        
        return False

    def retrieve_old_schedule(user, Shift):
        """Retrieves the user's previous schedule from the database"""
        shifts = Shift.objects.all().filter(user__exact=user.id)

        old_schedule = []

        for shift in shifts:
            old_schedule.append(RawShift(shift.shift_code, shift.date, ""))
        
        return old_schedule

    # Get shift codes/times for user
    shift_code_list = collect_shift_codes(user, ShiftCode)
    
    # Get all the stat holidays for the date range of the raw_schedule
    stat_holidays = collect_stat_holidays(raw_schedule, StatHoliday)

    # Get the default shift codes
    defaults = collect_defaults(config)
    
    # Assign start and end date/times to user's shifts
    schedule = []
    null_shifts = []
    missing_shifts = []
    
    for shift in raw_schedule:
        # Search for a shift match
        shift_match = False
            
        dow = shift.start_date.weekday()
        stat_match = is_stat(stat_holidays, shift.start_date)

        for code in shift_code_list:
	        # If matched, find the proper day to base shift details on
            
            if shift.shift_code == code.code:
                shift_match = True

                # Apply proper start time and duration
                if stat_match:
                    start_time = code.stat_start
                    duration = code.stat_duration
                elif dow == 0:
                    start_time = code.monday_start
                    duration = code.monday_duration
                elif dow == 1:
                    start_time = code.tuesday_start
                    duration = code.tuesday_duration
                elif dow == 2:
                    start_time = code.wednesday_start
                    duration = code.wednesday_duration
                elif dow == 3:
                    start_time = code.thursday_start
                    duration = code.thursday_duration
                elif dow == 4:
                    start_time = code.friday_start
                    duration = code.friday_duration
                elif dow == 5:
                    start_time = code.saturday_start
                    duration = code.saturday_duration
                elif dow == 6:
                    start_time = code.sunday_start
                    duration = code.sunday_duration

                if start_time:
                    # Shift has time, process as normal
                    
                    # Convert the decimal hours duration to h, m, and s
                    hours = int(duration)
                    minutes = int((duration*60) % 60)
                    seconds = int((minutes*60) % 60)
                
                    start_datetime = datetime.combine(shift.start_date, start_time)

                    end_datetime = start_datetime + timedelta(
                        hours=hours,
                        minutes=minutes,
                        seconds=seconds
                    )
  
                    schedule.append(FormattedShift(
                        shift.shift_code, start_datetime, end_datetime, 
                        shift.comment, code
                    ))
                else:
                    # Shift has no times - don't add to schedule and mark 
                    # it in the Null shift list
                    msg = "{} - {}".format(
                        get_formatted_date(shift.start_date), shift.shift_code
                    )

                    null_shifts.append(
                        EmailShift(shift.start_date, shift.shift_code, msg)
                    )

                # End loop
                break

        # If no shift match, provide default values
        if shift_match == False:
            # Add missing shift to the Missing shift list
            msg = "{} - {}".format(
                get_formatted_date(shift.start_date), shift.shift_code
            )

            missing_shifts.append(
                EmailShift(shift.start_date, shift.shift_code, msg)
            )
                
            # Set default times
            if stat_match:
                start_datetime = datetime.combine(
                    shift.start_date, defaults["stat_start"]
                )
                    
                end_datetime = start_datetime + timedelta(
                    hours=defaults["stat_hours"],
                    minutes=defaults["stat_minutes"],
                    seconds=defaults["stat_seconds"]
                )
            elif dow >= 5:
                start_datetime = datetime.combine(
                    shift.start_date, defaults["weekend_start"]
                )

                end_datetime = start_datetime + timedelta(
                    hours=defaults["weekend_hours"],
                    minutes=defaults["weekend_minutes"],
                    seconds=defaults["weekend_seconds"]
                )
            else:
                start_datetime = datetime.combine(
                    shift.start_date, defaults["weekday_start"]
                )
                    
                end_datetime = start_datetime + timedelta(
                    hours=defaults["weekday_hours"],
                    minutes=defaults["weekday_minutes"],
                    seconds=defaults["weekday_seconds"]
                )
  
            schedule.append(FormattedShift(
                shift.shift_code, start_datetime, end_datetime, 
                shift.comment, None
            ))

    
    # Determine the shift additions, deletions, and changes
    # Retrieve the old schedule
    old_schedule = retrieve_old_schedule(user, Shift)
    
    # Check if there are any deletions or changes
    deletions = []
    changes = []

    for old_shift in old_schedule:
        shift_delete = True
        shift_change = []
        changed_shifts = []

        for new_shift in schedule:
            if old_shift.start_date == new_shift.start_datetime.date():
                shift_delete = False

                if old_shift.shift_code == new_shift.shift_code:
                    shift_change.append(False)
                else:
                    shift_change.append(True)
                    changed_shifts.append(new_shift.shift_code)

        # Adds deletions
        if shift_delete == True:
            date = get_formatted_date(old_shift.start_date)
            msg = "%s - %s" % (get_formatted_date(old_shift.start_date), old_shift.shift_code)
            deletions.append(EmailShift(old_shift.start_date, old_shift.shift_code, msg))

        # Adds changes
        if len(shift_change) and False not in shift_change:
            shifts = "/".join(map(str, changed_shifts))
            msg = "%s - %s changed to %s" % (get_formatted_date(old_shift.start_date), old_shift.shift_code, shifts)
            changes.append(EmailShift(old_shift.start_date, shifts, msg))

    # Checks if there are any new additions
    additions = []

    for new_shift in schedule:
        shift_add = True

        for old_shift in old_schedule:
            if old_shift.start_date == new_shift.start_datetime.date():
                shift_add = False

        if shift_add == True:
            msg = "%s - %s" % (get_formatted_date(new_shift.start_datetime), new_shift.shift_code,)
            additions.append(EmailShift(new_shift.start_datetime, new_shift.shift_code, msg))

    # Checks if any of the null or missing shifts are new (i.e. haven't 
    # been reported to user yet)
    # Collects dates that could be new (additions or changes)
    mod_date = []

    for a in additions:
        mod_date.append(a.date)

    for d in deletions:
        mod_date.append(d.date)

    for c in changes:
        mod_date.append(c.date)

    # Removes any missing or null shifts not in the mod dates or are blank
    missing = []
    null = []

    for date in mod_date:
        for shift in missing_shifts:
            if shift.date == date and shift.shift.strip() != "":
                missing.append(shift)

        for shift in null_shifts:
            if shift.date == date and shift.shift.strip() != "":
                null.append(shift)

    return Schedule(schedule, additions, deletions, changes, missing, null)

def return_column_index(sheet, user, name_row, col_start, col_end):
    """Determines the Excel column containing the provided user"""
    index = None

    for i in range(col_start, col_end):
        try:
            if user.role == "p":
                cell_name = str(sheet.cell(row=name_row, column=i).value).strip()
            elif user.role == "a" or user.role == "t":
                cell_name = str(sheet.cell(name_row, i).value).strip()
            
            if cell_name.upper() == user.schedule_name.upper():
                index = i
                break
        except IndexError:
            break
        except Exception:
            log.exception("Error while searcing for column index for %s" 
                          % user.name)

    return index

def assemble_schedule(config, excel_files, user, ShiftCode, StatHoliday, Shift):
    # Setup the required Excel details
    role = user.role

    file_loc = excel_files[role]
    sheet = config.get("schedules", "sheet_{0}".format(role))
    name_row = config.getint("schedules", "name_row_{0}".format(role))
    col_start = config.getint("schedules", "name_col_start_{0}".format(role))
    col_end = config.getint("schedules", "name_col_end_{0}".format(role)) 
    row_start = config.getint("schedules", "shift_row_start_{0}".format(role))
    row_end = config.getint("schedules", "shift_row_end_{0}".format(role))
    date_col = config.getint("schedules", "date_col_{0}".format(role))

    # Open the proper Excel worksheet
    if user.role == "p":
        excel_book = openpyxl.load_workbook(file_loc)
        excel_sheet = excel_book[sheet]
    elif user.role == "a" or user.role == "t":
        excel_book = xlrd.open_workbook(file_loc)
        excel_sheet = excel_book.sheet_by_name(sheet)

    # Find column index for this user
    user_index = return_column_index(
        excel_sheet, user, name_row, col_start, col_end
    )

    # If the user.index is found, can run rest of program
    if user_index:
        raw_schedule = extract_raw_schedule(
            excel_book, excel_sheet, user, user_index, 
            row_start, row_end, date_col
        )

        formatted_schedule = generate_formatted_schedule(
            user, raw_schedule, ShiftCode, StatHoliday, config, Shift
        )

        return formatted_schedule

    else:
        log.warn(
            "Unable to find {0} (role = {1}) in the Excel schedule".format(
                user.name,
                user.role
            )
        )

        return None
